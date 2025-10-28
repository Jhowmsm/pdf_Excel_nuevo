from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Dict, List, Tuple
import fitz  # PyMuPDF
import uuid
import json
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = FastAPI(
    title="PDF Table Extractor",
    description="Extrae columnas estructuradas de tablas en PDF y las exporta a Excel (rango de páginas, franjas X).",
    version="1.3.0"
)

# ⚠ Ajusta estos dominios si tu codespace cambia de nombre
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://weary-cape-v6gp9gxg7jj52wqpw-8001.app.github.dev",  # backend
        "https://weary-cape-v6gp9gxg7jj52wqpw-5500.app.github.dev",  # frontend
        "http://localhost:5500",
        "http://127.0.0.1:5500",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
        "https://jhowmsm.github.io",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def normalizar(s: str) -> str:
    import unicodedata
    s = s.lower()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    s = s.replace(":", " ").replace("/", " ").replace(".", " ").replace(",", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def header_match(span_text: str, referencia: str) -> bool:
    norm_span = normalizar(span_text)
    norm_ref = normalizar(referencia)
    if norm_ref in norm_span or norm_span in norm_ref:
        return True
    palabras_ref = [p for p in norm_ref.split() if len(p) > 4]
    for p in palabras_ref:
        if p in norm_span:
            return True
    return False

# regex para NIF/NIE
PAT_ID = re.compile(
    r"""(
        [XYZ]\d{7}[A-Z]      # NIE tipo X/Y/Z + 7 dígitos + letra
        |
        \d{8}[A-Z]           # NIF 8 dígitos + letra
    )""",
    re.VERBOSE
)

@app.post("/procesar/")
async def procesar_pdf(
    file: UploadFile = File(...),
    referencias: str = Form(...),
    exclusiones: str = Form("{}"),
    pagina_inicio: int = Form(...),
    pagina_fin: int = Form(...)
):
    """
    parámetros:
    - referencias: lista JSON de headers esperados, en orden:
        ["NOMBRE Y APELLIDOS O RAZÓN SOCIAL", "NIF", "MARCA", "MATRÍCULA"]
    - exclusiones: dict JSON con textos a ignorar (típicamente los propios encabezados)
    - pagina_inicio / pagina_fin: páginas 1-based que queremos procesar
    """

    # Parsear referencias / exclusiones que vienen desde el frontend
    try:
        referencias_list: List[str] = json.loads(referencias)
        exclusiones_map: Dict[str, List[str]] = json.loads(exclusiones)
    except Exception:
        return {"error": "No se pudo leer referencias o exclusiones"}

    # Ajustar páginas a base 0
    start_page = max(pagina_inicio - 1, 0)
    end_page = max(pagina_fin - 1, start_page)

    original_name = file.filename or "archivo.pdf"
    temp_pdf = f"/tmp/{uuid.uuid4().hex}_{original_name}"
    with open(temp_pdf, "wb") as f:
        f.write(await file.read())

    doc = fitz.open(temp_pdf)
    end_page = min(end_page, len(doc) - 1)

    # Detectar NIF/NIE global solo en rango
    ids_detectados = set()
    for pageno in range(start_page, end_page + 1):
        page_text = doc[pageno].get_text("text")
        for m in PAT_ID.findall(page_text):
            ids_detectados.add(m)

    # === Parámetros de agrupación ===
    TOL_Y = 4  # tolerancia vertical para agrupar spans en una misma fila

    # Vamos a guardar todas las filas resultantes aquí:
    filas_totales: List[Dict[str, str]] = []

    # X de cada encabezado visto. Ej: {"NIF": 315.2, "MARCA": 400.1, ...}
    header_x_map: Dict[str, float] = {}

    # Bandera: no empezamos a capturar filas hasta ver headers reales dentro del rango
    en_modo_tabla = False

    # Recorremos SOLO el rango pedido
    for pageno in range(start_page, end_page + 1):
        page = doc[pageno]
        page_dict = page.get_text("dict")

        # 1. Detectar headers en esta página y guardar sus X
        pagina_tiene_header = False
        for block in page_dict["blocks"]:
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    texto_span = span["text"].strip()
                    x0, y0, x1, y1 = span["bbox"]

                    for ref in referencias_list:
                        if header_match(texto_span, ref):
                            pagina_tiene_header = True
                            # guardamos X si no estaba aún
                            if ref not in header_x_map:
                                header_x_map[ref] = x0

        if pagina_tiene_header:
            en_modo_tabla = True

        # si aún no hemos visto headers, no procesamos filas de esta página
        if not en_modo_tabla:
            continue

        # Asegurarnos de que tenemos las X críticas
        # Necesitamos al menos NIF, MARCA, MATRÍCULA
        # para poder dividir en franjas verticales
        # Si alguna falta, saltamos esta página
        header_keys_needed = ["NIF", "MARCA", "MATRÍCULA"]
        have_all = all(
            any(header_match(k, known_key) for k in header_x_map.keys())
            for known_key in header_keys_needed
        )
        # Nota: esto intenta ser flexible si cambia levemente "MATRÍCULA" vs "MATRICULA"
        # pero vamos a construir las X manualmente abajo de forma robusta

        # 2. Agrupar spans por "fila visual"
        filas_pag_cruda: List[Tuple[float, List[dict]]] = []
        # cada elemento será:
        #   (y_centro, [ { "text":..., "x0":..., "y0":..., ... }, ... ])

        for block in page_dict["blocks"]:
            for line in block.get("lines", []):
                spans = line.get("spans", [])
                if not spans:
                    continue

                # coordenada media vertical de esta línea
                ys = [s["bbox"][1] for s in spans] + [s["bbox"][3] for s in spans]
                y_centro = (min(ys) + max(ys)) / 2.0

                # buscar fila ya existente cercana en Y
                fila_idx = None
                for idx, (y_exist, _) in enumerate(filas_pag_cruda):
                    if abs(y_exist - y_centro) <= TOL_Y:
                        fila_idx = idx
                        break

                if fila_idx is None:
                    filas_pag_cruda.append((y_centro, []))
                    fila_idx = len(filas_pag_cruda) - 1

                # añadimos todos los spans de esta línea a esa "fila visual"
                for s in spans:
                    texto_span = s["text"].strip()
                    if not texto_span:
                        continue

                    # saltar numeritos sueltos tipo nº de página
                    if re.fullmatch(r"\d{1,3}", texto_span):
                        continue

                    # saltar encabezados repetidos
                    if any(header_match(texto_span, ref) for ref in referencias_list):
                        continue

                    # saltar textos explícitamente excluidos
                    skip_it = False
                    for ref in referencias_list:
                        if texto_span in exclusiones_map.get(ref, []):
                            skip_it = True
                            break
                    if skip_it:
                        continue

                    x0, y0, x1, y1 = s["bbox"]
                    filas_pag_cruda[fila_idx][1].append({
                        "text": texto_span,
                        "x0": x0,
                        "y0": y0,
                        "x1": x1,
                        "y1": y1
                    })

        # 3. Para poder cortar en franjas, necesitamos un mapa limpio de X:
        #    - x_nif
        #    - x_marca
        #    - x_matricula
        #    (la zona "nombre/apellidos/razón social" es todo lo que esté a la izquierda de x_nif)

        # Encontrar mejor aproximación de X para cada header clave, incluso si el texto varía
        def get_x_for(logical_name: str) -> float:
            # logical_name será "NIF", "MARCA", "MATRÍCULA"
            candidatos = []
            for ref_txt, x_value in header_x_map.items():
                if header_match(ref_txt, logical_name):
                    candidatos.append(x_value)
            if not candidatos:
                return None
            return min(candidatos)

        x_nif = get_x_for("NIF")
        x_marca = get_x_for("MARCA")
        x_matricula = get_x_for("MATRÍCULA") or get_x_for("MATRICULA")

        # si no tenemos estas X, nos es imposible segmentar esa página bien
        if x_nif is None or x_marca is None or x_matricula is None:
            # no agregamos filas de esta página porque no podemos separar columnas
            continue

        # 4. Ahora procesamos cada fila visual cruda y la convertimos en una fila lógica:
        for (_, spans_de_fila) in filas_pag_cruda:
            if not spans_de_fila:
                continue

            # Queremos armar:
            #   nombre_full = concat de spans con x0 < x_nif
            #   nif_val     = concat de spans x_nif <= x0 < x_marca
            #   marca_val   = concat de spans x_marca <= x0 < x_matricula
            #   matric_val  = concat de spans x0 >= x_matricula

            # Para la zona de nombre_full, además queremos respetar el orden izquierdo→derecho
            izquierda = []
            nif_chunks = []
            marca_chunks = []
            matr_chunks = []

            for sp in spans_de_fila:
                t = sp["text"]
                x0 = sp["x0"]

                # Clasificamos por franja X:
                if x0 < x_nif:
                    izquierda.append((x0, t))
                elif x0 < x_marca:
                    nif_chunks.append((x0, t))
                elif x0 < x_matricula:
                    marca_chunks.append((x0, t))
                else:
                    matr_chunks.append((x0, t))

            # Ordenamos cada lista por x0 ascendente para reconstruir en orden visual
            izquierda.sort(key=lambda z: z[0])
            nif_chunks.sort(key=lambda z: z[0])
            marca_chunks.sort(key=lambda z: z[0])
            matr_chunks.sort(key=lambda z: z[0])

            nombre_full = " ".join([txt for _, txt in izquierda]).strip()
            nif_val     = " ".join([txt for _, txt in nif_chunks]).strip()
            marca_val   = " ".join([txt for _, txt in marca_chunks]).strip()
            matr_val    = " ".join([txt for _, txt in matr_chunks]).strip()

            # descartamos filas totalmente vacías
            if not (nombre_full or nif_val or marca_val or matr_val):
                continue

            filas_totales.append({
                "NOMBRE": nombre_full,
                "NIF": nif_val,
                "MARCA": marca_val,
                "MATRICULA": matr_val
            })

    doc.close()

    # 5. Construir el Excel final
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Encabezados fijos en el orden que quieres entregar
    final_headers = [
        "NOMBRE Y APELLIDOS O RAZÓN SOCIAL",
        "NIF",
        "MARCA",
        "MATRÍCULA"
    ]
    ws.append(final_headers)

    # estilo de advertencia para nombres sospechosamente largos
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for fila in filas_totales:
        row_vals = [
            fila.get("NOMBRE", ""),
            fila.get("NIF", ""),
            fila.get("MARCA", ""),
            fila.get("MATRICULA", "")
        ]
        ws.append(row_vals)

        current_row_idx = ws.max_row
        nombre_val = ws.cell(current_row_idx, 1).value or ""
        if len(nombre_val) > 60:
            ws.cell(current_row_idx, 1).fill = amarillo

    # Hoja adicional con IDs detectados
    if ids_detectados:
        ws_alerta = wb.create_sheet("NIE_Warnings")
        ws_alerta.append(["NIF/NIE Detectados"])
        for codigo in sorted(ids_detectados):
            ws_alerta.append([codigo])

    output_path = f"/tmp/resultado_{uuid.uuid4().hex}.xlsx"
    wb.save(output_path)

    return FileResponse(
        output_path,
        filename="resultado.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
